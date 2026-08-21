import sys
import json
import openpyxl
import datetime
import re

def find_column_index(sheet, name_patterns, exclude_patterns=None):
    if isinstance(name_patterns, str):
        name_patterns = [name_patterns]
    if exclude_patterns is None:
        exclude_patterns = ["LAHIR", "BIRTH", "BORN"]
        
    first_row = list(sheet.iter_rows(max_row=1))[0]
    
    for pattern in name_patterns:
        clean_pattern = pattern.upper().replace(" ", "").replace("_", "").replace("/", "").replace(".", "")
        for c, cell in enumerate(first_row, 1):
            val = cell.value
            if val is not None:
                cleaned = str(val).strip().upper().replace(" ", "").replace("_", "").replace("/", "").replace(".", "")
                if any(ex in cleaned for ex in exclude_patterns):
                    continue
                if clean_pattern in cleaned:
                    return c
    return None

def parse_date_val(val_date):
    if val_date is None:
        return None
    if isinstance(val_date, (datetime.datetime, datetime.date)):
        return val_date
    if isinstance(val_date, (int, float)):
        try:
            return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(val_date))
        except ValueError:
            return None
    val_str = str(val_date).strip()
    if val_str == "":
        return None
    if re.match(r'^\d+(\.\d+)?$', val_str):
        try:
            return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(float(val_str)))
        except ValueError:
            return None
            
    formats = (
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y',
        '%d-%m-%y', '%d/%m/%y', '%d.%m.%y', '%y-%m-%d', '%y/%m/%d'
    )
    for fmt in formats:
        try:
            return datetime.datetime.strptime(val_str, fmt)
        except ValueError:
            pass
            
    m = re.search(r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', val_str)
    if m:
        try:
            return datetime.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
            
    m = re.search(r'(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})', val_str)
    if m:
        try:
            return datetime.datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    m = re.search(r'(\d{1,2})[-/.](\d{1,2})[-/.](\d{2})', val_str)
    if m:
        try:
            yr = int(m.group(3))
            yr = 2000 + yr if yr < 100 else yr
            return datetime.datetime(yr, int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    return None

def normalize_sales_date(val_date, target_year=None, target_month=None):
    if target_year is None:
        target_year = datetime.datetime.now().year
    if target_month is None:
        target_month = datetime.datetime.now().month

    dt = parse_date_val(val_date)
    if not dt:
        return None
    
    # Case 1: Already target_year and target_month
    if dt.year == target_year and dt.month == target_month:
        return dt
        
    # Case 2: Excel flipped month and day because of US locale (e.g. 2026-01-08 for 01/08/2026 -> 1 Aug 2026)
    if dt.year == target_year and dt.day == target_month and 1 <= dt.month <= 31:
        try:
            flipped_dt = datetime.datetime(dt.year, dt.day, dt.month)
            return flipped_dt
        except ValueError:
            pass

def extract_lm_day(val_date, target_year, target_month):
    dt = parse_date_val(val_date)
    if not dt:
        return None
        
    if dt.year == target_year and dt.month == target_month:
        return dt.day
        
    if dt.year == target_year and dt.day == target_month and dt.month <= 31:
        return dt.month

    if dt.year == target_year and (dt.month == target_month - 1 or dt.month == target_month) and 1 <= dt.day <= 31:
        return dt.day

    return None

def classify_stu_row(row_klasifikasi, row_series, row_type):
    combined = f"{row_klasifikasi or ''} {row_series or ''} {row_type or ''}".upper().strip()
    
    if 'OFF ROAD' in combined or 'OFFROAD' in combined or 'WR' in combined:
        cat = 'OFF ROAD'
        sname = 'WR SERIES'
    elif 'PREMIUM' in combined or 'NMAX' in combined or 'AEROX' in combined or 'XMAX' in combined or 'LEXI' in combined:
        cat = 'PREMIUM'
        if 'AEROX' in combined: sname = 'AEROX SERIES'
        elif 'XMAX' in combined: sname = 'XMAX SERIES'
        elif 'LEXI' in combined: sname = 'LEXI SERIES'
        else: sname = 'NMAX SERIES'
    elif 'ATM' in combined or 'GEAR' in combined or 'FREEGO' in combined:
        cat = 'ATM'
        if 'FREEGO' in combined: sname = 'FREEGO SERIES'
        else: sname = 'GEAR SERIES'
    elif 'CLASSY' in combined or 'FAZZIO' in combined or 'FILANO' in combined or 'GRAND' in combined:
        cat = 'CLASSY'
        if 'FILANO' in combined or 'GRAND' in combined: sname = 'FILANO SERIES'
        else: sname = 'FAZZIO SERIES'
    elif 'MOPED' in combined or 'MX' in combined or 'JUPITER' in combined or 'VEGA' in combined:
        cat = 'MOPED'
        if 'JUPITER' in combined: sname = 'JUPITER SERIES'
        elif 'VEGA' in combined: sname = 'VEGA SERIES'
        else: sname = 'MX SERIES'
    elif 'SPORT' in combined or 'R15' in combined or 'XSR' in combined or 'VIXION' in combined:
        cat = 'SPORT'
        if 'R15' in combined: sname = 'R15 SERIES'
        elif 'XSR' in combined: sname = 'XSR SERIES'
        elif 'VIXION' in combined: sname = 'VIXION SERIES'
        else: sname = 'SPORT SERIES'
    elif 'AT STD' in combined or 'MIO' in combined or 'XRIDE' in combined or 'X-RIDE' in combined:
        cat = 'AT STD'
        if 'XRIDE' in combined or 'X-RIDE' in combined: sname = 'X-RIDE SERIES'
        else: sname = 'MIO SERIES'
    else:
        cat = 'PREMIUM'
        sname = 'NMAX SERIES'
        
    return cat, sname

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided"}))
        sys.exit(1)
        
    file_path = sys.argv[1]
    mode = sys.argv[2] if len(sys.argv) > 2 else "both"
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        acv = 0
        lm = 0
        lm_full = 0
        stock_2024 = 0
        stock_2025 = 0
        stock_2026 = 0
        stock_breakdown = {}
        stu_breakdown = {}
        pos_breakdown = {}
        l_cash = 0
        l_kredit = 0
        fincoy_ach = {"ADIRA": 0, "BAF": 0, "IMFI": 0, "MEGA": 0, "SOF": 0}
        daily_counts = {}
        max_day_found = 0
        
        # 1. Parse ACV
        if mode in ("both", "acv"):
            # Find Sales Sheet
            sheet_stu = None
            for name in wb.sheetnames:
                if "STU" in name.upper() or "SALES" in name.upper() or "JUAL" in name.upper():
                    sheet_stu = wb[name]
                    break
            if not sheet_stu:
                if "Sheet1" in wb.sheetnames:
                    sheet_stu = wb["Sheet1"]
                elif len(wb.sheetnames) > 0:
                    sheet_stu = wb[wb.sheetnames[0]]
                    
            if sheet_stu:
                col_aa = find_column_index(sheet_stu, ["RANGKA", "MESIN"])
                col_no = find_column_index(sheet_stu, ["NO", "NOMOR", "SPK"])
                col_typ = find_column_index(sheet_stu, ["TYPE", "TIPE", "MOTOR"])
                col_consumer = find_column_index(sheet_stu, ["KONSUMEN", "PELANGGAN", "BUYER", "PEMBELI"])
                col_pos = find_column_index(sheet_stu, ["DEALER", "POS", "PENJUAL"])
                col_ser = find_column_index(sheet_stu, ["SERIES"])
                col_klas = find_column_index(sheet_stu, ["KLASIFIKASI", "KATEGORI"])
                col_leasing = find_column_index(sheet_stu, ["LEASING", "FINCOY", "FINANCE", "LSG", "BEBAN", "CARABAYAR", "CARA BAYAR", "PAYMENT", "SYSTEM", "SISTEM"])

                # Priority date search: TGL JUAL first, then TGL DEL/DO
                col_date = find_column_index(sheet_stu, ["TGLJUAL", "TANGGALJUAL", "TGLJUALJL"], exclude_patterns=["LAHIR", "BIRTH", "BORN"])
                if not col_date:
                    col_date = find_column_index(sheet_stu, ["TGLDEL", "TANGGALDEL", "TGLDO", "TANGGALDO", "DEL"], exclude_patterns=["LAHIR", "BIRTH", "BORN"])
                if not col_date:
                    col_date = find_column_index(sheet_stu, ["TGL", "TANGGAL", "DATE"], exclude_patterns=["LAHIR", "BIRTH", "BORN"])

                l_cash = 0
                l_kredit = 0
                fincoy_ach = {"ADIRA": 0, "BAF": 0, "IMFI": 0, "MEGA": 0, "SOF": 0}
                pos_breakdown = {}

                target_year = datetime.datetime.now().year
                target_month = datetime.datetime.now().month

                for row in sheet_stu.iter_rows(min_row=2, max_row=1000):
                    val_aa = str(row[col_aa - 1].value or "").strip() if col_aa and col_aa <= len(row) else ""
                    val_consumer = str(row[col_consumer - 1].value or "").strip() if col_consumer and col_consumer <= len(row) else ""
                    val_leasing_raw = str(row[col_leasing - 1].value or "").strip() if col_leasing and col_leasing <= len(row) else ""

                    has_real_aa = val_aa not in ("", "-", "NONE", "#N/A")
                    has_real_consumer = val_consumer not in ("", "-", "NONE", "#N/A")
                    has_real_leasing = val_leasing_raw not in ("", "-", "NONE", "#N/A")

                    # Row is valid if it has real chassis/engine OR customer name OR leasing method
                    if has_real_aa or has_real_consumer or has_real_leasing:
                        acv += 1

                        # Date processing & daily performance
                        dt = None
                        if col_date and col_date <= len(row):
                            val_date = row[col_date - 1].value
                            dt = normalize_sales_date(val_date, target_year, target_month)

                        if dt:
                            day_num = dt.day
                        else:
                            # Fallback day number extraction if val_date is numeric or float string
                            day_num = None
                            if col_date and col_date <= len(row):
                                raw_v = str(row[col_date - 1].value or "").strip()
                                if raw_v.replace('.', '', 1).isdigit():
                                    try:
                                        dn = int(float(raw_v))
                                        if 1 <= dn <= 31:
                                            day_num = dn
                                    except ValueError:
                                        pass
                            if not day_num:
                                day_num = max_day_found if max_day_found > 0 else 1

                        if 1 <= day_num <= 31:
                            daily_counts[day_num] = daily_counts.get(day_num, 0) + 1
                            if day_num > max_day_found:
                                max_day_found = day_num

                        # Parse STU breakdown
                        val_klas = row[col_klas - 1].value if col_klas and col_klas <= len(row) else None
                        val_ser = row[col_ser - 1].value if col_ser and col_ser <= len(row) else None
                        val_typ = row[col_typ - 1].value if col_typ and col_typ <= len(row) else None
                        cat_name, series_name = classify_stu_row(val_klas, val_ser, val_typ)

                        # Track STU breakdown
                        if cat_name not in stu_breakdown:
                            stu_breakdown[cat_name] = {}
                        stu_breakdown[cat_name][series_name] = stu_breakdown[cat_name].get(series_name, 0) + 1

                        # Track ATM CLASSY program sales
                        if cat_name in ('ATM', 'CLASSY'):
                            if 'program_atm_classy' not in stu_breakdown:
                                stu_breakdown['program_atm_classy'] = {
                                    'acv': 0, 'cash': 0, 'credit': 0,
                                    'adira': 0, 'imfi': 0, 'sof': 0, 'baf': 0, 'mega': 0
                                }
                            p_data = stu_breakdown['program_atm_classy']
                            p_data['acv'] += 1
                            val_p_upper = val_leasing_raw.upper()
                            if val_p_upper in ("", "-", "NONE", "CASH", "TUNAI", "DIRECT", "CASH P", "CASH S", "KDS"):
                                p_data['cash'] += 1
                            else:
                                p_data['credit'] += 1
                                if "ADIRA" in val_p_upper:
                                    p_data['adira'] += 1
                                elif "IMFI" in val_p_upper or "INDOMOBIL" in val_p_upper:
                                    p_data['imfi'] += 1
                                elif "SOF" in val_p_upper or "SUMMIT" in val_p_upper or "OTO" in val_p_upper:
                                    p_data['sof'] += 1
                                elif "BAF" in val_p_upper or "BUSSAN" in val_p_upper:
                                    p_data['baf'] += 1
                                elif "MEGA" in val_p_upper or "MAF" in val_p_upper or val_p_upper == "MF":
                                    p_data['mega'] += 1

                        # Parse POS breakdown (Rule: SO1, SO 1, SO-1, RIAU 1, D1... -> DEALER)
                        val_pos_raw = str(row[col_pos - 1].value or "").upper().strip() if col_pos and col_pos <= len(row) else ""
                        matched_pos = 'DEALER'
                        if any(term in val_pos_raw for term in ["SO1", "SO 1", "SO-1", "RIAU 1"]) or val_pos_raw in ("", "-", "NONE", "DEALER") or val_pos_raw.startswith("D1"):
                            matched_pos = 'DEALER'
                        else:
                            for sub_p in ["RIAU 2", "GARUDA SAKTI", "LIPAT KAIN", "BANGKINANG", "PETAPAHAN", "BELILAS", "PERANAP", "SUNGAI LALA", "PADANG LUAS", "KERINCI", "UKUI", "PERAWANG", "KERINCI KANAN", "SIAK", "SABAK AUH", "PANCUR BATU"]:
                                if sub_p in val_pos_raw:
                                    matched_pos = sub_p
                                    break

                        pos_breakdown[matched_pos] = pos_breakdown.get(matched_pos, 0) + 1

                        # Parse LEASING column from branch spreadsheet
                        val_p = val_leasing_raw.upper()
                        if val_p in ("", "-", "NONE", "CASH", "TUNAI", "DIRECT", "CASH P", "CASH S", "KDS"):
                            l_cash += 1
                        else:
                            l_kredit += 1
                            if "ADIRA" in val_p:
                                fincoy_ach["ADIRA"] = fincoy_ach.get("ADIRA", 0) + 1
                            elif "BAF" in val_p or "BUSSAN" in val_p:
                                fincoy_ach["BAF"] = fincoy_ach.get("BAF", 0) + 1
                            elif "IMFI" in val_p or "INDOMOBIL" in val_p:
                                fincoy_ach["IMFI"] = fincoy_ach.get("IMFI", 0) + 1
                            elif "MEGA" in val_p or "MAF" in val_p or val_p == "MF":
                                fincoy_ach["MEGA"] = fincoy_ach.get("MEGA", 0) + 1
                            elif "SOF" in val_p or "SUMMIT" in val_p or "OTO" in val_p:
                                fincoy_ach["SOF"] = fincoy_ach.get("SOF", 0) + 1
                            else:
                                clean_name = val_p.replace("PT", "").replace(".", "").strip()
                                if clean_name:
                                    fincoy_ach[clean_name] = fincoy_ach.get(clean_name, 0) + 1

        daily_performance = []
        if max_day_found == 0 and acv > 0:
            max_day_found = 1
            daily_counts[1] = acv

        if max_day_found > 0:
            cum_sum = 0
            for day in range(1, 32):
                if day <= max_day_found:
                    cum_sum += daily_counts.get(day, 0)
                    if day == max_day_found and cum_sum < acv:
                        cum_sum = acv
                    daily_performance.append(cum_sum)
                else:
                    daily_performance.append(None)
        else:
            daily_performance = [None] * 31

        # 2. Parse LM
        if mode in ("both", "lm"):
            # Find LM Sheet
            sheet_lm = None
            for name in wb.sheetnames:
                if "LM" in name.upper() or "LAST MONTH" in name.upper() or "LASTMONTH" in name.upper():
                    sheet_lm = wb[name]
                    break
            if not sheet_lm and mode == "lm":
                # Fallback to STU
                for name in wb.sheetnames:
                    if "STU" in name.upper() or "SALES" in name.upper() or "JUAL" in name.upper():
                        sheet_lm = wb[name]
                        break
                if not sheet_lm:
                    if "Sheet1" in wb.sheetnames:
                        sheet_lm = wb["Sheet1"]
                    elif len(wb.sheetnames) > 0:
                        sheet_lm = wb[wb.sheetnames[0]]
                        
            if sheet_lm:
                now = datetime.datetime.now()
                current_year = now.year
                current_month = now.month
                if current_month == 1:
                    lm_year = current_year - 1
                    lm_month = 12
                else:
                    lm_year = current_year
                    lm_month = current_month - 1

                today_day = now.day

                col_aa = find_column_index(sheet_lm, ["RANGKA", "MESIN"]) or 1
                col_date = find_column_index(sheet_lm, ["TGLJUAL", "TANGGALJUAL", "TGLDEL", "TANGGALDEL", "TGL", "TANGGAL", "DATE"]) or 6

                for row in sheet_lm.iter_rows(min_row=2, max_row=2000):
                    if col_aa <= len(row):
                        val_aa = row[col_aa - 1].value
                        if val_aa is not None and str(val_aa).strip() != "":
                            val_str = str(val_aa).strip().upper()
                            if not val_str.startswith("TOTAL") and not val_str.startswith("JUMLAH") and not val_str.startswith("AVERAGE"):
                                val_date = row[col_date - 1].value if col_date <= len(row) else None
                                day = extract_lm_day(val_date, lm_year, lm_month)
                                if day is not None:
                                    lm_full += 1
                                    if 1 <= day <= today_day:
                                        lm += 1
                                else:
                                    if "LM" in sheet_lm.title.upper() or "LAST" in sheet_lm.title.upper() or "JUL" in sheet_lm.title.upper():
                                        lm_full += 1

        # 3. Parse Stock
        if mode in ("both", "acv"):
            # Find Stock Sheet
            sheet_stok = None
            for name in wb.sheetnames:
                if "STOK" in name.upper() or "STOCK" in name.upper():
                    sheet_stok = wb[name]
                    break
            # We process stock sheet if found
            stock_breakdown = {}
            if sheet_stok:
                col_year = find_column_index(sheet_stok, "TAHUNRAKITAN") or find_column_index(sheet_stok, "TAHUN") or 16
                col_class = find_column_index(sheet_stok, "KLASIFIKASITIPE") or find_column_index(sheet_stok, "KLASIFIKASI") or 1
                for row in sheet_stok.iter_rows(min_row=2):
                    if col_year <= len(row):
                        val = row[col_year - 1].value
                        if val is not None:
                            try:
                                year = int(float(val))
                                if year == 2024:
                                    stock_2024 += 1
                                elif year == 2025:
                                    stock_2025 += 1
                                elif year == 2026:
                                    stock_2026 += 1
                            except ValueError:
                                pass
                    if col_class <= len(row):
                        val_class = row[col_class - 1].value
                        if val_class is not None:
                            class_name = str(val_class).strip().upper()
                            if class_name != "" and not class_name.startswith("TOTAL") and not class_name.startswith("JUMLAH"):
                                stock_breakdown[class_name] = stock_breakdown.get(class_name, 0) + 1
                                
        wb.close()
        act_ytd_jan_2026 = lm_full + acv
        print(json.dumps({
            "acv": acv,
            "lm": lm,
            "lm_full": lm_full,
            "act_ytd_jan_2026": act_ytd_jan_2026,
            "stock_2024": stock_2024,
            "stock_2025": stock_2025,
            "stock_2026": stock_2026,
            "stock_breakdown": stock_breakdown,
            "stu_breakdown": stu_breakdown,
            "pos_breakdown": pos_breakdown,
            "leasing_breakdown": {"cash": l_cash, "kredit": l_kredit, "fincoy": fincoy_ach},
            "daily_performance": daily_performance
        }))
        
    except Exception as e:
        if 'wb' in locals():
            try:
                wb.close()
            except:
                pass
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()
